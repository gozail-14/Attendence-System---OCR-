import re
import time
import requests
import openpyxl
from datetime import datetime
import cv2
from PIL import Image
import pytesseract

# ------------------ CONFIG ------------------
AZURE_ENDPOINT = "https://attendence-system1.cognitiveservices.azure.com/"
AZURE_KEY = "G77m6Y0ciEurfsJsJoWHSTHcEb3vfU09Z4O7gwW8285vymKAptpHJQQJ99BDAC8vTInXJ3w3AAAFACOG7bZ2"
EXCEL_FILE = "2025_Spring_List.xlsx"
ATTENDANCE_IMAGE = "attendance.jpg"
# --------------------------------------------

def preprocess_image(image_path):
    """Enhance image for better OCR results."""
    img = cv2.imread(image_path)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gray, (3,3), 0)
    _, thresh = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    processed_path = "processed_attendance.jpg"
    cv2.imwrite(processed_path, thresh)
    return processed_path

def extract_numbers_with_tesseract(image_path):
    """Fallback OCR using Tesseract for handwritten numbers."""
    text = pytesseract.image_to_string(Image.open(image_path))
    numbers = set(re.findall(r"\d{5,}", text))
    return numbers

def extract_numbers_from_image(image_path):
    """
    Extract numbers from image using Azure OCR with fallback to Tesseract.
    """
    print("📌 Preprocessing image...")
    processed_image = preprocess_image(image_path)

    ocr_url = AZURE_ENDPOINT + "vision/v3.2/read/analyze"
    headers = {
        "Ocp-Apim-Subscription-Key": AZURE_KEY,
        "Content-Type": "application/octet-stream"
    }

    try:
        with open(processed_image, "rb") as img_file:
            response = requests.post(ocr_url, headers=headers, data=img_file)
        response.raise_for_status()
        operation_url = response.headers["Operation-Location"]

        # Poll Azure OCR
        while True:
            result = requests.get(operation_url, headers={"Ocp-Apim-Subscription-Key": AZURE_KEY})
            result.raise_for_status()
            result_json = result.json()
            if result_json["status"] == "succeeded":
                break
            elif result_json["status"] == "failed":
                raise Exception("Azure OCR failed")
            time.sleep(1)

        # Extract numbers from Azure OCR
        text_lines = []
        for read_result in result_json["analyzeResult"]["readResults"]:
            for line in read_result["lines"]:
                text_lines.append(line["text"])

        print("\n📌 RAW OCR TEXT LINES FROM AZURE:")
        for line in text_lines:
            print(line)

        numbers = set()
        for line in text_lines:
            found_nums = re.findall(r"\d{5,}", line)
            numbers.update(found_nums)

        # If Azure fails or returns few numbers, use Tesseract
        if len(numbers) < 3:
            print("\n⚠️ Azure OCR may have missed IDs. Using Tesseract fallback...")
            numbers.update(extract_numbers_with_tesseract(processed_image))

    except Exception as e:
        print(f"\n❌ Azure OCR failed: {e}. Using Tesseract fallback...")
        numbers = extract_numbers_with_tesseract(processed_image)

    print("\n✅ Extracted Student IDs:", numbers)
    return numbers

def mark_attendance(excel_file, present_ids):
    """Marks attendance in Excel file."""
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active

    date_str = input("Enter date (e.g., 2025-09-10): ").strip()
    batch_str = input("Enter batch name (e.g., Batch A): ").strip()
    header_name = f"{date_str}_{batch_str}"

    # Find Std Nbr column
    std_nbr_col = None
    for col in range(1, sheet.max_column + 1):
        if str(sheet.cell(row=1, column=col).value).strip().lower() == "std nbr":
            std_nbr_col = col
            break
    if not std_nbr_col:
        print("❌ 'Std Nbr' column not found!")
        return

    # Next empty column
    next_col = sheet.max_column + 1
    sheet.cell(row=1, column=next_col, value=header_name)

    print("\n📌 Student IDs in Excel:")
    for row in range(2, sheet.max_row + 1):
        student_id = str(sheet.cell(row=row, column=std_nbr_col).value).strip()
        print(student_id, end=" | ")
        if student_id in present_ids:
            sheet.cell(row=row, column=next_col, value="Present")
        else:
            sheet.cell(row=row, column=next_col, value="")

    wb.save(excel_file)
    print(f"\n✅ Attendance updated in {excel_file} under column '{header_name}'")

if __name__ == "__main__":
    present_ids = extract_numbers_from_image(ATTENDANCE_IMAGE)
    mark_attendance(EXCEL_FILE, present_ids)
