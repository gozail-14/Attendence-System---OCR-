from flask import Flask, request, render_template, jsonify
import os
import json
import re
import time
import requests
import openpyxl
from datetime import datetime
import cv2
from PIL import Image
import pytesseract

app = Flask(__name__)
UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# Azure Config
AZURE_ENDPOINT = "https://attendence-system1.cognitiveservices.azure.com/"
AZURE_KEY = "G77m6Y0ciEurfsJsJoWHSTHcEb3vfU09Z4O7gwW8285vymKAptpHJQQJ99BDAC8vTInXJ3w3AAAFACOG7bZ2"
EXCEL_FILE = "2025_Spring_List.xlsx"

def preprocess_image(image_path, output_folder="uploads"):
    img = cv2.imread(image_path)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gray, (3,3), 0)
    _, thresh = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    processed_path = os.path.join(output_folder, "bw_" + os.path.basename(image_path))
    cv2.imwrite(processed_path, thresh)
    return processed_path

def extract_numbers_with_tesseract(image_path):
    text = pytesseract.image_to_string(Image.open(image_path))
    numbers = set(re.findall(r"\d{5,}", text))
    return numbers

def extract_numbers_from_image(image_path):
    # Always use the processed (BW) image for OCR
    processed_image = preprocess_image(image_path, output_folder=app.config["UPLOAD_FOLDER"])
    ocr_url = AZURE_ENDPOINT + "vision/v3.2/read/analyze"
    headers = {
        "Ocp-Apim-Subscription-Key": AZURE_KEY,
        "Content-Type": "application/octet-stream"
    }

    try:
        with open(processed_image, "rb") as img_file:
            response = requests.post(ocr_url, headers=headers, data=img_file)
        response.raise_for_status()
        operation_url = response.headers["Operation-Location"]

        while True:
            result = requests.get(operation_url, headers={"Ocp-Apim-Subscription-Key": AZURE_KEY})
            result.raise_for_status()
            result_json = result.json()
            if result_json["status"] == "succeeded":
                break
            elif result_json["status"] == "failed":
                raise Exception("Azure OCR failed")
            time.sleep(1)

        numbers = set()
        for read_result in result_json["analyzeResult"]["readResults"]:
            for line in read_result["lines"]:
                found_nums = re.findall(r"\d{5,}", line["text"])
                numbers.update(found_nums)

        if len(numbers) < 3:
            numbers.update(extract_numbers_with_tesseract(processed_image))

    except Exception as e:
        numbers = extract_numbers_with_tesseract(processed_image)

    return numbers

def get_student_data():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    
    students = []
    std_nbr_col = None
    name_col = None
    
    for col in range(1, sheet.max_column + 1):
        header = str(sheet.cell(row=1, column=col).value).strip().lower()
        if header == "std nbr":
            std_nbr_col = col
        elif "name" in header:
            name_col = col
    
    if std_nbr_col:
        for row in range(2, sheet.max_row + 1):
            student_id = str(sheet.cell(row=row, column=std_nbr_col).value).strip()
            name = str(sheet.cell(row=row, column=name_col).value).strip() if name_col else "Unknown"
            if student_id and student_id != "None":
                students.append({"id": student_id, "name": name})
    
    return students

def mark_attendance_in_excel(present_ids, date_str, batch_str):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    
    header_name = f"{date_str}_{batch_str}"
    
    std_nbr_col = None
    for col in range(1, sheet.max_column + 1):
        if str(sheet.cell(row=1, column=col).value).strip().lower() == "std nbr":
            std_nbr_col = col
            break
    
    if not std_nbr_col:
        return False
    
    next_col = sheet.max_column + 1
    sheet.cell(row=1, column=next_col, value=header_name)
    
    for row in range(2, sheet.max_row + 1):
        student_id = str(sheet.cell(row=row, column=std_nbr_col).value).strip()
        if student_id in present_ids:
            sheet.cell(row=row, column=next_col, value="Present")
        else:
            sheet.cell(row=row, column=next_col, value="")
    
    wb.save(EXCEL_FILE)
    return True

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/upload", methods=["POST"])
def upload_file():
    try:
        file = request.files["image"]
        date_str = request.form.get("date", datetime.now().strftime("%Y-%m-%d"))
        batch_str = request.form.get("batch", "BatchA")
        excel_file = request.files.get("excel")

        if not file:
            return jsonify({"error": "No image file uploaded"}), 400

        filepath = os.path.join(app.config["UPLOAD_FOLDER"], file.filename)
        file.save(filepath)

        # Use uploaded Excel file if provided
        if excel_file:
            excel_filename = os.path.join(app.config["UPLOAD_FOLDER"], excel_file.filename)
            excel_file.save(excel_filename)
            excel_path = excel_filename
        else:
            excel_path = EXCEL_FILE

        present_ids = extract_numbers_from_image(preprocess_image(filepath, output_folder=app.config["UPLOAD_FOLDER"]))

        # Pass excel_path to attendance functions
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active

        # Mark attendance in selected Excel file
        header_name = f"{date_str}_{batch_str}"
        std_nbr_col = None
        for col in range(1, sheet.max_column + 1):
            if str(sheet.cell(row=1, column=col).value).strip().lower() == "std nbr":
                std_nbr_col = col
                break
        if not std_nbr_col:
            return jsonify({"error": "'Std Nbr' column not found!"}), 400

        next_col = sheet.max_column + 1
        sheet.cell(row=1, column=next_col, value=header_name)
        for row in range(2, sheet.max_row + 1):
            student_id = str(sheet.cell(row=row, column=std_nbr_col).value).strip()
            if student_id in present_ids:
                sheet.cell(row=row, column=next_col, value="Present")
            else:
                sheet.cell(row=row, column=next_col, value="")
        wb.save(excel_path)

        # Get student data from selected Excel file
        students = []
        name_col = None
        for col in range(1, sheet.max_column + 1):
            header = str(sheet.cell(row=1, column=col).value).strip().lower()
            if header == "std nbr":
                std_nbr_col = col
            elif "name" in header:
                name_col = col
        if std_nbr_col:
            for row in range(2, sheet.max_row + 1):
                student_id = str(sheet.cell(row=row, column=std_nbr_col).value).strip()
                name = str(sheet.cell(row=row, column=name_col).value).strip() if name_col else "Unknown"
                if student_id and student_id != "None":
                    students.append({"id": student_id, "name": name})

        present_students = [s for s in students if s["id"] in present_ids]

        return jsonify({
            "success": True,
            "present_count": len(present_students),
            "present_students": present_students,
            "total_students": len(students)
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    app.run(debug=True)
